from attr import attr
import pandas as pd
from geopy.geocoders import Nominatim
from openpyxl import load_workbook
import translators as ts
import pycountry

# a class containing all data wrangling functions
class fun:
    def __init__(self) -> None:
        pass
    # remove not translatable colors
    def changeStr(self,rrr):
        if "mét." in rrr:
            return rrr[:-5]
        else:
            return rrr

    # simplify body type name
    def changeBody(self,b):
        if " / " in b:
            return b.split(" / ")[0]
        else:
            return b
    # remove pandas formatting from excel
    def removeFormatting(self,ws):
        # ws is not the worksheet name, but the worksheet object
        for row in ws.iter_rows():
            for cell in row:
                cell.style = 'Normal'
    # translate to english german content
    def getEnglish(self,lst):
        di = {'Name': lst}
        vls = list()
        try:
            for v in di.values():
                for i in v:
                    if i != "null":
                        vls.append(ts.google(i, from_language='de', to_language='en'))
                    else:
                        vls.append('Other')
            di2 = {'Values':vls}
            di.update(di2)
            return di
        except:
            return di
    # obtain country code based on city name
    def getCountry(self,lst):
        di = {'Name': lst}
        vls = list()
        geolocator = Nominatim(user_agent="MyApp")
        try:
            for v in di.values():
                for i in v:
                    if i != "null":
                        location = geolocator.geocode(i, language="en") 
                        ctyname = list(location.address.split(", "))[-1]
                        ctycode = pycountry.countries.get(name=ctyname).alpha_2
                        vls.append(ctycode)
                    else:
                        vls.append('Other')
            di2 = {'Values':vls}
            di.update(di2)
            return di
        except:
            return di
    # derive mileage consumption unit
    def consUnit(self,txt):
        if 'l/100km' in txt:
            return 'l_km_consumption'
        else:
            return 'null'

    # derive mileage unit    
    def mil(self,txt):
        if 'km' in txt:
            return 'kilometer'
        else:
            return 'mile'
    # derive Left Hand Drive - LHD
    # or Right Hand Drive from Country Code
    def drive(self,txt):
        if 'GB' in txt:
            return 'RHD'
        else:
            return 'LHD'

    # format mileage in the required format
    def mileage_conv(self,val):
        val = int(val)
        return '{0:.2f}'.format(val)

# a class containing data rework functions        
class datarework:
    def __init__(self) -> None:
        pass

    def read_json(self, data):
        return pd.read_json(data, lines=True)

    def pivotdata(self,data):
        return data.pivot(index='ID', columns='Attribute Names', values='Attribute Values').reset_index()

    def dropdupl(self,data):
        return data[['ID','MakeText','TypeName','TypeNameFull','ModelText','ModelTypeText']].drop_duplicates()

    def mergedata(self,dr,dl):
        return pd.merge(dr,dl,how='inner',on='ID')

    def enrichdata(self,dr,dl,coln,ncoln):
        norm = pd.merge(dr, dl, how='left', left_on=[coln], right_on=['Name'])
        norm.drop(columns={'Name',coln}, inplace=True)
        norm.rename(columns={'Values':ncoln}, inplace=True)
        return norm

# a class containing export data functions
class exportdata:
    def __init__(self) -> None:
        pass
    
    def export2excel(self,template,df1,df2):
        try:
            book = load_workbook(template)
            writer = pd.ExcelWriter(template, engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

            df1.to_excel(writer, "preprocessing", index=False)
            df2.to_excel(writer, "normalisation", index=False)
            df2.to_excel(writer, "integration", index=False, header=False, startrow=writer.sheets['integration'].max_row)

            for ws in writer.sheets.values():
                f.removeFormatting(ws)

            writer.save()
        except:
            return print('Export failed.')

# import classes
f = fun()
dw = datarework()
ed = exportdata()

# import data
rawdata = dw.read_json('supplier_car.json')
attributes = dw.pivotdata(rawdata)
rawu = dw.dropdupl(rawdata)
preprodata = dw.mergedata(rawu,attributes)

ex_preprodata = preprodata.sort_values('ID').reset_index(drop=True)
normalized = ex_preprodata.copy()

# keep only needed columns
normalized = normalized[
    [
        'MakeText',
        'ModelText',
        'ModelTypeText',
        'BodyColorText',
        'BodyTypeText',
        'City',
        'ConditionTypeText',
        'ConsumptionTotalText',
        'FirstRegMonth',
        'FirstRegYear',
        'Km'
    ]
]
# fill null values with Other
normalized = normalized.fillna('Other')

# format Body Color and Body Type
normalized.BodyColorText = normalized.BodyColorText.apply(f.changeStr)
normalized.BodyTypeText = normalized.BodyTypeText.apply(f.changeBody)

# obtain list of unique values for colors, body types, condition and cities
colors = list(normalized.BodyColorText.drop_duplicates())
colors = [ x for x in colors if "Mét." not in x ]
body = list(normalized.BodyTypeText.drop_duplicates().dropna())
condition = list(normalized.ConditionTypeText.drop_duplicates())
cities = list(normalized.City.drop_duplicates())

# leverage external APIs to obtain data for missing features in input data
print("Getting additional data via APIs...")
bodies = pd.DataFrame.from_dict(f.getEnglish(body))
conditions = pd.DataFrame.from_dict(f.getEnglish(condition))
cols = pd.DataFrame.from_dict(f.getEnglish(colors))
cts = pd.DataFrame.from_dict(f.getCountry(cities))

# enriching normalized dataframe with new features
normalized = dw.enrichdata(normalized,bodies,'BodyTypeText','carType')
normalized = dw.enrichdata(normalized,conditions,'ConditionTypeText','condition')
normalized = dw.enrichdata(normalized,cols,'BodyColorText','color')

# adding country code
normalized = pd.merge(normalized, cts, how='left', left_on=['City'], right_on=['Name'])
normalized.drop(columns={'Name'}, inplace=True)
normalized.rename(columns={'Values':'country'}, inplace=True)

# creating new columns based on target data
normalized['fuel_consumption_unit'] = normalized.ConsumptionTotalText.apply(f.consUnit)
normalized['mileage_unit'] = normalized.ConsumptionTotalText.apply(f.mil)
normalized.drop(columns={'ConsumptionTotalText'},inplace=True)
normalized['drive'] = normalized.country.apply(f.drive)

normalized.rename(columns={
    'MakeText':'make',
    'ModelText':'model',
    'ModelTypeText':'model_variant',
    'City':'city',
    'FirstRegMonth':'manufacture_month',
    'FirstRegYear':'manufacture_year',
    'Km':'mileage'
},inplace=True)

# adding missing columns required in target data 
normalized['currency'] = "CHF"
normalized['price_on_request'] = "false"
normalized['type'] = "car"
normalized['zip'] = "null"

# formatting columns based on target data
normalized.make = normalized.make.apply(lambda str : str.title())
normalized.color = normalized.color.apply(lambda str : str.title())
normalized.mileage = normalized.mileage.apply(f.mileage_conv)

# reorder columns
newcols = [
    'carType',
    'color',
    'condition',
    'currency',
    'drive',
    'city',
    'country',
    'make',
    'manufacture_year',
    'mileage',
    'mileage_unit',
    'model',
    'model_variant',
    'price_on_request',
    'type',
    'zip',
    'manufacture_month',
    'fuel_consumption_unit'
]
normalized = normalized.reindex(columns=newcols)

# exporting data to excel
print("Attempting to export data to excel...")
ed.export2excel('extemp.xlsx',ex_preprodata,normalized)

print("Script terminated.")